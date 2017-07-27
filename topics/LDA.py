


from gensim import corpora, models
import openpyxl as xls
import sys
from pprint import pprint
import numpy as np
import gc



wb = xls.load_workbook("pre/clinton.xlsx",read_only=True)
print wb.get_sheet_names()
ws = wb.get_sheet_by_name('Sheet 1')



n = 1 + 3171 # num of rows in your excel
# n = 1 + 8906


row_ranges = range(1, n) # excel index starts from 1

for i, row in enumerate(row_ranges):
    row_str = 'A' + str(row)
    row_ranges[i] = row_str

# print row_ranges


docs = []
for row in row_ranges:

    cell = ws[row]
    if cell.value == 'x':
        continue
    docs.append(cell.value)


# print len(docs)


wb.close()
gc.collect()


stoplist = set(u''.split())
texts = [[word for word in document.lower().split() if word not in stoplist ] for document in docs]

# remove words that appear only once
from collections import defaultdict
frequency = defaultdict(int)
for text in texts:
     for token in text:
         frequency[token] += 1

texts = [[token for token in text if frequency[token] > 1]
         for text in texts]

# from pprint import pprint  # pretty-printer
# pprint(texts)

print len(texts)



for num_of_topics in [6,8,10,12]:


    dictionary = corpora.Dictionary(texts)
    corpus = [dictionary.doc2bow(text) for text in texts]
    ldamodel = models.ldamodel.LdaModel(corpus, num_topics=num_of_topics , id2word = dictionary,random_state=2046, passes = 2, chunksize = 800, update_every = 1)


    doc_lda = ldamodel[corpus]

    # doc_lda[0] # probabilities of each doc belong to a topic

    gammaDF = [[0 for _ in xrange(num_of_topics)] for _ in xrange(len(doc_lda))]
    for i,x in enumerate(doc_lda):
        for j in range(len(x)):
            idx = x[j][0]
            gammaDF[i][idx] = x[j][1]

    # a = [[4,3,2],[0,1,5]]
    # zip(*a)

    n_row = len(gammaDF)
    percentage = [sum(x)/n_row  for x in zip(*gammaDF)]

    # percentage


    topics = ldamodel.show_topics(num_topics = num_of_topics , num_words=20, formatted=False)

    print len(topics)

    res = [list(zip(*topic[1])[0]) for topic in topics]


    for i, x in enumerate(percentage):
        res[i].insert(0, 'Topic '+str(i+1))
        res[i].append("%.2f" % round(x*100,2)+"%")




    dest_filename = 'res/clinton_' + str(num_of_topics) + 'topics.xlsx'
    wb2 = xls.Workbook()
    ws2 = wb2.active
    ws2.title = 'topics'

    for line in zip(*res):
        ws2.append(line)

    wb2.save(filename = dest_filename)
